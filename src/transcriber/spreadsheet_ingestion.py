"""Spreadsheet ingestion utilities for scorecards (XLSX/CSV)."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Any, Dict, List, Optional


@dataclass
class ScorecardTable:
    name: str
    headers: List[str]
    rows: List[Dict[str, Any]]


def _sniff_csv(data: bytes) -> ScorecardTable:
    text = data.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader]
    headers = list(reader.fieldnames or [])
    return ScorecardTable(name="uploaded.csv", headers=headers, rows=rows)


def _load_xlsx(data: bytes) -> ScorecardTable:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to process XLSX files") from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: List[Dict[str, Any]] = []
    headers: List[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        values = ["" if v is None else v for v in row]
        if i == 0:
            headers = [str(h).strip() for h in values]
            continue
        record = {str(headers[j] if j < len(headers) else f"col{j}"): values[j] for j in range(len(values))}
        rows.append(record)
    return ScorecardTable(name=ws.title or "Sheet1", headers=headers, rows=rows)


def load_scorecard(uploaded_file) -> ScorecardTable:
    name = uploaded_file.name
    raw = uploaded_file.read()
    uploaded_file.seek(0)
    suffix = name.lower().rsplit(".", 1)[-1] if "." in name else ""

    if suffix in {"csv"}:
        return _sniff_csv(raw)
    if suffix in {"xlsx", "xlsm"}:
        return _load_xlsx(raw)

    # Fallback: attempt CSV
    return _sniff_csv(raw)
